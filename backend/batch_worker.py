import sys
import os
from openpyxl import Workbook, load_workbook

# --------------------------------------------------
# IMPORT YOUR EXISTING EXTRACTION FUNCTION
# --------------------------------------------------
from worker import extract_invoice_data  
# extract_invoice_data(pdf_path)
# returns: (invoice_no_dt, buyer_address, invoice_value, exchange_rate)

# --------------------------------------------------
# Arguments
# --------------------------------------------------
if len(sys.argv) < 3:
    print("Usage: python batch_worker.py output.xlsx pdf1.pdf pdf2.pdf ...")
    sys.exit(1)

output_excel = sys.argv[1]
pdf_files = sys.argv[2:]

# --------------------------------------------------
# Create or load Excel
# --------------------------------------------------
if os.path.exists(output_excel):
    wb = load_workbook(output_excel)
    ws = wb.active
    start_row = ws.max_row + 1
else:
    wb = Workbook()
    ws = wb.active

    # HEADER ROW (WRITE ONCE)
    ws.append([
        "S.No",
        "Invoice No & Date",
        "Buyer Name & Address",
        "Invoice Value",
        "Exchange Rate"
    ])
    start_row = 2

# --------------------------------------------------
# Process each PDF
# --------------------------------------------------
serial_no = start_row - 1

for pdf_path in pdf_files:
    try:
        print(f"Processing: {pdf_path}")

        invoice_no_dt, buyer, value, rate = extract_invoice_data(pdf_path)

        serial_no += 1

        ws.append([
            serial_no,
            invoice_no_dt,
            buyer,
            value,
            rate,
        ])

    except Exception as e:
        # 🔥 DO NOT CRASH THE BATCH
        print(f"ERROR processing {pdf_path}: {e}")
        continue

# --------------------------------------------------
# Save Excel
# --------------------------------------------------
wb.save(output_excel)

print(f"Batch Excel created: {output_excel}")
